import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import datetime

def generate_excel_report(test_results, output_path="cybershield_test_analysis.xlsx"):
    """
    Generates a beautifully styled, premium Excel analysis report from test results.
    
    test_results is a list of dicts:
    {
        "id": "test_web_functional_dashboard_loads",
        "name": "Verify the main dashboard loads successfully.",
        "platform": "Web",
        "category": "Functional",
        "status": "PASS", # PASS, FAIL, SKIP
        "duration": 0.05,
        "error": ""
    }
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Harmonious Palette (Sleek Dark Theme accents for headers, clean fills)
    PRIMARY_COLOR = "4F46E5"    # Indigo (Primary)
    ACCENT_COLOR = "06B6D4"     # Cyan (Accent)
    PASS_COLOR = "10B981"       # Emerald Green
    FAIL_COLOR = "EF4444"       # Rose Red
    SKIP_COLOR = "F59E0B"       # Amber Orange
    HEADER_FILL = "1F2937"      # Dark Gray (Header)
    ZEBRA_FILL = "F8FAFC"       # Very Light Gray/Blue
    CARD_BG = "EEF2F6"          # Soft Gray
    
    # Fonts
    font_title = Font(name="Segoe UI", size=18, bold=True, color="FFFFFF")
    font_section = Font(name="Segoe UI", size=14, bold=True, color="1F2937")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=11, color="1F2937")
    font_bold = Font(name="Segoe UI", size=11, bold=True, color="1F2937")
    
    # Fills
    fill_header = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    fill_title = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
    fill_zebra = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")
    fill_card = PatternFill(start_color=CARD_BG, end_color=CARD_BG, fill_type="solid")
    
    fill_pass = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    fill_fail = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    fill_skip = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    
    font_pass = Font(name="Segoe UI", size=11, bold=True, color="065F46")
    font_fail = Font(name="Segoe UI", size=11, bold=True, color="991B1B")
    font_skip = Font(name="Segoe UI", size=11, bold=True, color="92400E")

    # Borders
    thin_border_side = Side(border_style="thin", color="CBD5E1")
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom = Border(bottom=Side(border_style="medium", color="1F2937"))
    
    # Convert list to DataFrame for aggregations
    df = pd.DataFrame(test_results)
    
    # ----------------------------------------------------
    # SHEET 1: DASHBOARD
    # ----------------------------------------------------
    ws_dash = wb.create_sheet(title="Dashboard")
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Banner Header
    ws_dash.merge_cells("A1:F2")
    title_cell = ws_dash["A1"]
    title_cell.value = "  CyberShield Automation E2E Testing Dashboard"
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = Alignment(vertical="center", indent=1)
    
    # Subtitle info
    ws_dash["A3"] = "Generated At:"
    ws_dash["A3"].font = font_bold
    ws_dash["B3"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_dash["B3"].font = font_body
    
    ws_dash["E3"] = "Total Test Cases:"
    ws_dash["E3"].font = font_bold
    ws_dash["F3"] = len(df)
    ws_dash["F3"].font = font_body
    
    # Summary Card blocks
    total = len(df)
    passed = len(df[df["status"] == "PASS"])
    failed = len(df[df["status"] == "FAIL"])
    skipped = len(df[df["status"] == "SKIP"])
    pass_rate = (passed / total * 100) if total > 0 else 0
    
    # Place Metric cards
    metrics = [
        ("Total Scanned", total, "A5:B6", PRIMARY_COLOR),
        ("Passed Tests", passed, "C5:C6", PASS_COLOR),
        ("Failed Tests", failed, "D5:D6", FAIL_COLOR),
        ("Skipped Tests", skipped, "E5:E6", SKIP_COLOR),
        ("Pass Rate", f"{pass_rate:.1f}%", "F5:F6", ACCENT_COLOR)
    ]
    
    # Helper to style metric card cells
    for label, val, range_str, accent in metrics:
        ws_dash.merge_cells(range_str)
        top_left = range_str.split(":")[0]
        ws_dash[top_left] = f"{label}\n{val}"
        ws_dash[top_left].font = Font(name="Segoe UI", size=12, bold=True, color="1F2937")
        ws_dash[top_left].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_dash[top_left].fill = fill_card
        
        # Draw borders around merged block
        for row in range_str_to_cells(range_str):
            for c in row:
                ws_dash[c].border = border_all
                
    # Platform breakdown table
    ws_dash["A8"] = "Platform Summary"
    ws_dash["A8"].font = font_section
    
    ws_dash["A9"] = "Platform"
    ws_dash["B9"] = "Scanned"
    ws_dash["C9"] = "Passed"
    ws_dash["D9"] = "Failed"
    ws_dash["E9"] = "Skipped"
    ws_dash["F9"] = "Pass Rate"
    
    for col_let in ["A", "B", "C", "D", "E", "F"]:
        ws_dash[f"{col_let}9"].font = font_header
        ws_dash[f"{col_let}9"].fill = fill_header
        ws_dash[f"{col_let}9"].alignment = Alignment(horizontal="center")
        
    platforms = ["Web", "Mobile"]
    row_idx = 10
    for p in platforms:
        p_df = df[df["platform"] == p]
        p_total = len(p_df)
        p_passed = len(p_df[p_df["status"] == "PASS"])
        p_failed = len(p_df[p_df["status"] == "FAIL"])
        p_skipped = len(p_df[p_df["status"] == "SKIP"])
        p_rate = (p_passed / p_total * 100) if p_total > 0 else 0
        
        ws_dash[f"A{row_idx}"] = p
        ws_dash[f"B{row_idx}"] = p_total
        ws_dash[f"C{row_idx}"] = p_passed
        ws_dash[f"D{row_idx}"] = p_failed
        ws_dash[f"E{row_idx}"] = p_skipped
        ws_dash[f"F{row_idx}"] = f"{p_rate:.1f}%"
        
        for col_let in ["A", "B", "C", "D", "E", "F"]:
            cell = ws_dash[f"{col_let}{row_idx}"]
            cell.font = font_body
            cell.border = border_all
            if col_let != "A":
                cell.alignment = Alignment(horizontal="center")
        row_idx += 1
        
    # Category breakdown table
    ws_dash[f"A{row_idx+1}"] = "Testing Category Breakdown"
    ws_dash[f"A{row_idx+1}"].font = font_section
    
    cat_start = row_idx + 2
    ws_dash[f"A{cat_start}"] = "Category"
    ws_dash[f"B{cat_start}"] = "Scanned"
    ws_dash[f"C{cat_start}"] = "Passed"
    ws_dash[f"D{cat_start}"] = "Failed"
    ws_dash[f"E{cat_start}"] = "Skipped"
    ws_dash[f"F{cat_start}"] = "Pass Rate"
    
    for col_let in ["A", "B", "C", "D", "E", "F"]:
        ws_dash[f"{col_let}{cat_start}"].font = font_header
        ws_dash[f"{col_let}{cat_start}"].fill = fill_header
        ws_dash[f"{col_let}{cat_start}"].alignment = Alignment(horizontal="center")
        
    cats = sorted(df["category"].unique())
    c_row = cat_start + 1
    for c in cats:
        c_df = df[df["category"] == c]
        c_total = len(c_df)
        c_passed = len(c_df[c_df["status"] == "PASS"])
        c_failed = len(c_df[c_df["status"] == "FAIL"])
        c_skipped = len(c_df[c_df["status"] == "SKIP"])
        c_rate = (c_passed / c_total * 100) if c_total > 0 else 0
        
        ws_dash[f"A{c_row}"] = c
        ws_dash[f"B{c_row}"] = c_total
        ws_dash[f"C{c_row}"] = c_passed
        ws_dash[f"D{c_row}"] = c_failed
        ws_dash[f"E{c_row}"] = c_skipped
        ws_dash[f"F{c_row}"] = f"{c_rate:.1f}%"
        
        for col_let in ["A", "B", "C", "D", "E", "F"]:
            cell = ws_dash[f"{col_let}{c_row}"]
            cell.font = font_body
            cell.border = border_all
            if col_let != "A":
                cell.alignment = Alignment(horizontal="center")
            if c_row % 2 == 1:
                cell.fill = fill_zebra
        c_row += 1

    # ----------------------------------------------------
    # SHEET 2: DETAILED RESULTS
    # ----------------------------------------------------
    ws_details = wb.create_sheet(title="Detailed Results")
    ws_details.views.sheetView[0].showGridLines = True
    
    headers = ["Test ID", "Platform", "Category", "Test Case Description", "Status", "Duration (s)", "Assertion Failure Log"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_details.cell(row=1, column=col_idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(border_style="medium", color="1F2937"))
        
    for r_idx, t in enumerate(test_results, 2):
        ws_details.cell(row=r_idx, column=1, value=t["id"]).font = font_body
        ws_details.cell(row=r_idx, column=2, value=t["platform"]).font = font_body
        ws_details.cell(row=r_idx, column=3, value=t["category"]).font = font_body
        ws_details.cell(row=r_idx, column=4, value=t["name"]).font = font_body
        
        status_cell = ws_details.cell(row=r_idx, column=5, value=t["status"])
        if t["status"] == "PASS":
            status_cell.fill = fill_pass
            status_cell.font = font_pass
        elif t["status"] == "FAIL":
            status_cell.fill = fill_fail
            status_cell.font = font_fail
        else:
            status_cell.fill = fill_skip
            status_cell.font = font_skip
        status_cell.alignment = Alignment(horizontal="center")
            
        dur_cell = ws_details.cell(row=r_idx, column=6, value=t["duration"])
        dur_cell.font = font_body
        dur_cell.alignment = Alignment(horizontal="center")
        
        ws_details.cell(row=r_idx, column=7, value=t["error"]).font = font_body
        
        # Apply borders and zebra striping
        for col_idx in range(1, 8):
            c = ws_details.cell(row=r_idx, column=col_idx)
            c.border = border_all
            if r_idx % 2 == 1 and col_idx != 5:
                c.fill = fill_zebra

    # ----------------------------------------------------
    # SHEET 3: SECURITY & UX REVIEWS
    # ----------------------------------------------------
    ws_reviews = wb.create_sheet(title="Analysis & Recommendations")
    ws_reviews.views.sheetView[0].showGridLines = True
    
    ws_reviews.merge_cells("A1:D2")
    rev_title = ws_reviews["A1"]
    rev_title.value = "  AI Security & Compliance Recommendations"
    rev_title.font = font_title
    rev_title.fill = PatternFill(start_color=ACCENT_COLOR, end_color=ACCENT_COLOR, fill_type="solid")
    rev_title.alignment = Alignment(vertical="center", indent=1)
    
    # Headers
    ws_reviews["A4"] = "Impact Area"
    ws_reviews["B4"] = "Compliance Standard"
    ws_reviews["C4"] = "Risk Level"
    ws_reviews["D4"] = "Remediation Strategy"
    
    for col_let in ["A", "B", "C", "D"]:
        ws_reviews[f"{col_let}4"].font = font_header
        ws_reviews[f"{col_let}4"].fill = fill_header
        ws_reviews[f"{col_let}4"].alignment = Alignment(horizontal="center")
        
    recommendations = [
        ("Security", "OWASP MASVS / Mobile Injection Checks", "HIGH", 
         "Enforce parameterized input sanitization on all custom Compose text fields and check secure SQLite database keys."),
        ("Performance", "DOM Latency / Rendering Thresholds", "MEDIUM",
         "Implement local tokenizer thread offloading using Kotlin Coroutines in BullyingClassifier to keep Compose UI at 60fps."),
        ("Accessibility", "WCAG 2.1 Contrast & Touch Target Rules", "MEDIUM",
         "Modify custom visual elements to ensure a minimum size of 48x48dp and check color combinations for WCAG AA compliance."),
        ("Security", "OAuth Credentials Encryption", "CRITICAL",
         "Migrate standard user tokens from shared preferences to EncryptedSharedPreferences for standard Android secure storage."),
        ("Mobile-Specific", "SMS Alert Dispatch Buffer", "HIGH",
         "Validate dispatch filters to throttle notifications. Ensure background threads do not run indefinitely in listeners.")
    ]
    
    r_row = 5
    for area, std, risk, strategy in recommendations:
        ws_reviews[f"A{r_row}"] = area
        ws_reviews[f"B{r_row}"] = std
        
        risk_cell = ws_reviews[f"C{r_row}"]
        risk_cell.value = risk
        risk_cell.alignment = Alignment(horizontal="center")
        if risk in ["HIGH", "CRITICAL"]:
            risk_cell.fill = fill_fail
            risk_cell.font = font_fail
        else:
            risk_cell.fill = fill_skip
            risk_cell.font = font_skip
            
        ws_reviews[f"D{r_row}"] = strategy
        
        for col_let in ["A", "B", "C", "D"]:
            cell = ws_reviews[f"{col_let}{r_row}"]
            cell.font = font_body
            cell.border = border_all
            if col_let == "D":
                cell.alignment = Alignment(wrap_text=True)
            if r_row % 2 == 1 and col_let != "C":
                cell.fill = fill_zebra
        r_row += 1

    # Auto-adjust column widths across all sheets
    for ws in [ws_dash, ws_details, ws_reviews]:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    lines = str(cell.value).split('\n')
                    for line in lines:
                        if len(line) > max_len:
                            max_len = len(line)
            col_letter = get_column_letter(col[0].column)
            # Constrain widths to reasonable sizing
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)
            
    # Specific adjustment for details description column
    ws_details.column_dimensions["D"].width = 40
    ws_details.column_dimensions["G"].width = 35
    ws_reviews.column_dimensions["D"].width = 45
            
    wb.save(output_path)
    print(f"Excel Report Successfully generated at: {output_path}")

def range_str_to_cells(range_str):
    """Parses range strings like A5:F6 and returns list of lists of cell names."""
    start, end = range_str.split(":")
    start_col, start_row = "", ""
    for char in start:
        if char.isalpha():
            start_col += char
        else:
            start_row += char
            
    end_col, end_row = "", ""
    for char in end:
        if char.isalpha():
            end_col += char
        else:
            end_row += char
            
    s_col_idx = openpyxl.utils.column_index_from_string(start_col)
    e_col_idx = openpyxl.utils.column_index_from_string(end_col)
    
    rows = []
    for r in range(int(start_row), int(end_row) + 1):
        col_list = []
        for c in range(s_col_idx, e_col_idx + 1):
            col_let = get_column_letter(c)
            col_list.append(f"{col_let}{r}")
        rows.append(col_list)
    return rows

if __name__ == "__main__":
    # Test generation with sample list
    samples = [
        {"id": "test_web_func_1", "name": "Load page", "platform": "Web", "category": "Functional", "status": "PASS", "duration": 0.05, "error": ""},
        {"id": "test_web_sec_2", "name": "SQL check", "platform": "Web", "category": "Security", "status": "FAIL", "duration": 0.08, "error": "AssertionError: query not sanitized"},
        {"id": "test_mobile_spec_3", "name": "SMS test", "platform": "Mobile", "category": "Mobile-Specific", "status": "SKIP", "duration": 0.01, "error": "No SMS card inserted"}
    ]
    generate_excel_report(samples)
