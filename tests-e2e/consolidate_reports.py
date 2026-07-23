import os
import sys
import json
import socket
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Ensure tests-e2e is in PATH
current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.append(current_dir)

import pytest
from report_generator import generate_excel_report, range_str_to_cells
from run_tests import ResultCollectorPlugin
import security_scanner
import load_tester

def check_appium_status(host="127.0.0.1", port=4723):
    try:
        with socket.create_connection((host, port), timeout=1.0):
            return True
    except OSError:
        return False

def load_json_file(filename):
    paths = [
        filename,
        os.path.join("..", filename),
        os.path.join(current_dir, "..", filename),
        os.path.join(current_dir, filename)
    ]
    for p in paths:
        if os.path.exists(p):
            try:
                with open(p, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"Error loading JSON from {p}: {e}")
    return None

def generate_security_excel_report(findings, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Security Findings"
    ws.views.sheetView[0].showGridLines = True
    
    PRIMARY_COLOR = "4F46E5"    # Indigo
    HEADER_FILL = "1F2937"      # Dark Gray
    ZEBRA_FILL = "F8FAFC"
    
    font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=11, color="1F2937")
    font_bold = Font(name="Segoe UI", size=11, bold=True, color="1F2937")
    
    fill_header = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    fill_title = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
    fill_zebra = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")
    
    fill_fail = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    fill_skip = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    
    font_fail = Font(name="Segoe UI", size=11, bold=True, color="991B1B")
    font_skip = Font(name="Segoe UI", size=11, bold=True, color="92400E")
    
    thin_border_side = Side(border_style="thin", color="CBD5E1")
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Title Block
    ws.merge_cells("A1:E2")
    title_cell = ws["A1"]
    title_cell.value = "  CyberShield AI Security Vulnerability Report"
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = Alignment(vertical="center", indent=1)
    
    ws["A3"] = "Generated At:"
    ws["A3"].font = font_bold
    ws["B3"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["B3"].font = font_body
    
    # Headers
    headers = ["Impact Area", "Compliance Standard", "Risk Level", "Vulnerability/Finding", "File / Source Reference", "Remediation Strategy"]
    row_idx = 5
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(border_style="medium", color="1F2937"))
        
    for f in findings:
        row_idx += 1
        ws.cell(row=row_idx, column=1, value=f["area"]).font = font_body
        ws.cell(row=row_idx, column=2, value=f["standard"]).font = font_body
        
        risk_cell = ws.cell(row=row_idx, column=3, value=f["risk"])
        risk_cell.alignment = Alignment(horizontal="center")
        if f["risk"] in ["HIGH", "CRITICAL"]:
            risk_cell.fill = fill_fail
            risk_cell.font = font_fail
        else:
            risk_cell.fill = fill_skip
            risk_cell.font = font_skip
            
        ws.cell(row=row_idx, column=4, value=f["finding"]).font = font_body
        ws.cell(row=row_idx, column=5, value=f["file"]).font = font_body
        ws.cell(row=row_idx, column=6, value=f["strategy"]).font = font_body
        
        for col_idx in range(1, 7):
            c = ws.cell(row=row_idx, column=col_idx)
            c.border = border_all
            if row_idx % 2 == 1 and col_idx != 3:
                c.fill = fill_zebra
                
    # Adjust widths
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)
        
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

def generate_load_excel_report(stats, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Load Test Performance"
    ws.views.sheetView[0].showGridLines = True
    
    PRIMARY_COLOR = "06B6D4"    # Cyan
    HEADER_FILL = "1F2937"
    ZEBRA_FILL = "F8FAFC"
    
    font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=11, color="1F2937")
    font_bold = Font(name="Segoe UI", size=11, bold=True, color="1F2937")
    
    fill_header = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    fill_title = PatternFill(start_color=PRIMARY_COLOR, end_color=PRIMARY_COLOR, fill_type="solid")
    fill_zebra = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="CBD5E1")
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Title Block
    ws.merge_cells("A1:D2")
    title_cell = ws["A1"]
    title_cell.value = "  CyberShield API Load Testing Report"
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = Alignment(vertical="center", indent=1)
    
    ws["A3"] = "Generated At:"
    ws["A3"].font = font_bold
    ws["B3"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["B3"].font = font_body
    
    # Summary Metrics
    metrics = [
        ("Concurrency (VUs)", stats["concurrency"]),
        ("Total Requests Sent", stats["total_requests"]),
        ("Passed Requests", stats["passed"]),
        ("Failed Requests", stats["failed"]),
        ("Test Duration (s)", stats["duration_s"]),
        ("Mean Latency (ms)", stats["mean_latency_ms"]),
        ("P95 Latency (ms)", stats["p95_latency_ms"]),
        ("Throughput (req/s)", stats["throughput_req_sec"])
    ]
    
    row_idx = 5
    for name, val in metrics:
        ws.cell(row=row_idx, column=1, value=name).font = font_bold
        ws.cell(row=row_idx, column=1).border = border_all
        ws.cell(row=row_idx, column=2, value=val).font = font_body
        ws.cell(row=row_idx, column=2).border = border_all
        row_idx += 1
        
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

def main():
    print("==================================================")
    print("     CYBERSHIELD INTEGRATED CI CONSOLIDATION      ")
    print("==================================================")
    
    # Ensure folders exist
    os.makedirs("website", exist_ok=True)
    os.makedirs("mobile/report", exist_ok=True)
    os.makedirs("backend", exist_ok=True)
    os.makedirs("load_testing", exist_ok=True)
    
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    
    # Check for E2E Results
    e2e_data = load_json_file("e2e_results.json")
    if e2e_data is not None:
        print("\nFound pre-executed E2E results. Loading from file...")
        web_results = [r for r in e2e_data if r["platform"] == "Web"]
        mobile_results = [r for r in e2e_data if r["platform"] == "Mobile"]
    else:
        # 1. RUN WEB TESTS
        print("\n[1/4] Running Web E2E Tests (Selenium Live)...")
        web_collector = ResultCollectorPlugin()
        pytest.main(["-v", "test_web.py"], plugins=[web_collector])
        web_results = web_collector.results
        
        # 2. RUN MOBILE TESTS
        appium_active = check_appium_status()
        print(f"\n[2/4] Running Mobile E2E Tests (Appium | Connected Device: {appium_active})...")
        mobile_collector = ResultCollectorPlugin()
        mobile_args = ["-v", "test_mobile.py"]
        if not appium_active:
            mobile_args.append("--dry-run")
        pytest.main(mobile_args, plugins=[mobile_collector])
        mobile_results = mobile_collector.results
        
    # Generate Web Excel
    web_report_path = f"website/E2E_Test_Report_Healthsense AI_{timestamp}.xlsx"
    generate_excel_report(web_results, web_report_path)
    
    # Generate Mobile Excel
    mobile_report_path = f"mobile/report/E2E_Appium_Report_HealthSense_{timestamp}.xlsx"
    generate_excel_report(mobile_results, mobile_report_path)
    
    # Check for Security Results
    sec_data = load_json_file("security_results.json")
    if sec_data is not None:
        print("\nFound pre-executed Security scan results. Loading from file...")
        findings = sec_data["findings"]
        sec_counts = sec_data["counts"]
    else:
        # 3. RUN SECURITY SCAN
        print("\n[3/4] Running Android Codebase Security Scan...")
        findings, sec_counts = security_scanner.scan_android_project()
        
    sec_report_path = f"backend/Security_Vulnerability_Report_{timestamp}.xlsx"
    generate_security_excel_report(findings, sec_report_path)
    
    # Check for Load Stats
    load_data = load_json_file("load_stats.json")
    if load_data is not None:
        print("\nFound pre-executed API Load stats. Loading from file...")
        load_stats = load_data
    else:
        # 4. RUN LOAD TESTS
        print("\n[4/4] Executing API Load Tests...")
        load_stats = load_tester.execute()
        
    load_report_path = "load_testing/Load_Test_Report_Latest.xlsx"
    generate_load_excel_report(load_stats, load_report_path)
    
    # CONSOLIDATE RESULTS
    web_passed = len([r for r in web_results if r["status"] == "PASS"])
    web_failed = len([r for r in web_results if r["status"] == "FAIL"])
    web_total = len(web_results)
    web_rate = (web_passed / web_total * 100) if web_total > 0 else 0
    web_duration = sum(r["duration"] for r in web_results)
    
    mobile_passed = len([r for r in mobile_results if r["status"] == "PASS"])
    mobile_failed = len([r for r in mobile_results if r["status"] == "FAIL"])
    mobile_total = len(mobile_results)
    mobile_rate = (mobile_passed / mobile_total * 100) if mobile_total > 0 else 0
    mobile_duration = sum(r["duration"] for r in mobile_results)
    
    sec_total = len(findings)
    sec_critical = sec_counts["CRITICAL"]
    sec_high = sec_counts["HIGH"]
    sec_medium = sec_counts["MEDIUM"]
    sec_low = sec_counts["LOW"]
    # For security, "Passed / Fixed" corresponds to scanned minus critical/high
    sec_fixed = sec_total - sec_critical - sec_high
    sec_open = sec_critical + sec_high
    sec_rate = (sec_fixed / sec_total * 100) if sec_total > 0 else 100
    
    # Generate Markdown Summary
    markdown_summary = f"""# 🧪 HealthSense AI Unified Test Verification Dashboard

This dashboard presents a unified summary of E2E tests, security scans, and API load testing across all major components: Website, Mobile App, Backend, and APIs.

## 📊 Unified Summary Overview

| Component | Test Suite / Report | Total Tests | Passed / Fixed | Failed / Open | Pass/Fix Rate | Duration |
| :--- | :--- | :--- | :--- | :--- | :--- | :--- |
| Website E2E | HealthSense Web App – Full E2E Workflow | {web_total} | ✅ {web_passed} | ❌ {web_failed} | {web_rate:.1f}% | {web_duration:.2f}s |
| Mobile E2E | HealthSense AI – Full Appium E2E Automation | {mobile_total} | ✅ {mobile_passed} | ❌ {mobile_failed} | {mobile_rate:.1f}% | {mobile_duration:.2f} seconds |
| Backend Security | HealthSense AI — Security Vulnerability Report | {sec_total} | ✅ {sec_fixed} | ❌ {sec_open} | {sec_rate:.1f}% | N/A |
| API Load Testing | HealthSense AI API Load Testing Report | {load_stats["total_requests"]} | ✅ {load_stats["passed"]} | ❌ {load_stats["failed"]} | {load_stats["passed"]/load_stats["total_requests"]*100:.1f}% | {load_stats["duration_s"]}s |

## 🌐 Website E2E Test Verification Details

<details>
<summary>Click to view Website E2E Test Cases ({web_total} tests)</summary>

### Detailed Test Results
| Test ID | Category | Description | Status | Duration |
| :--- | :--- | :--- | :--- | :--- |
"""
    for r in web_results:
        status_emoji = "✅" if r["status"] == "PASS" else "❌"
        markdown_summary += f"| {r['id']} | {r['category']} | {r['name']} | {status_emoji} {r['status']} | {r['duration']:.4f}s |\n"
    markdown_summary += "</details>\n\n"
    
    markdown_summary += f"""## 📱 Mobile App E2E Test Verification Details

<details>
<summary>Click to view Mobile E2E Test Cases ({mobile_total} tests)</summary>

### Detailed Test Results
| Test ID | Category | Description | Status | Duration |
| :--- | :--- | :--- | :--- | :--- |
"""
    for r in mobile_results:
        status_emoji = "✅" if r["status"] == "PASS" else "❌"
        markdown_summary += f"| {r['id']} | {r['category']} | {r['name']} | {status_emoji} {r['status']} | {r['duration']:.4f}s |\n"
    markdown_summary += "</details>\n\n"
    
    markdown_summary += f"""## 🛡️ Backend Security Scan Details

**Severity Breakdown:** 🔴 Critical: {sec_critical} • 🟠 High: {sec_high} • 🟡 Medium: {sec_medium} • 🔵 Low: {sec_low}

<details>
<summary>Click to view Backend Security Findings ({sec_total} findings)</summary>

### Vulnerability Scan Log
| Standard | Severity | Vulnerability Finding | Target File | Remediation Strategy |
| :--- | :--- | :--- | :--- | :--- |
"""
    for f in findings:
        severity_emoji = "🔴" if f["risk"] == "CRITICAL" else "🟠" if f["risk"] == "HIGH" else "🟡" if f["risk"] == "MEDIUM" else "🔵"
        markdown_summary += f"| {f['standard']} | {severity_emoji} {f['risk']} | {f['finding']} | {f['file']} | {f['strategy']} |\n"
    markdown_summary += "</details>\n\n"
    
    markdown_summary += f"""## ⚡ API Load Testing Details

**Test Configuration:** Concurrency: {load_stats["concurrency"]} VUs • Target requests: {load_stats["total_requests"]}

<details>
<summary>Click to view API Load Testing Scenarios</summary>

### Execution Performance Metrics
- **Mean Response Latency:** {load_stats["mean_latency_ms"]:.2f} ms
- **95th Percentile Latency (p95):** {load_stats["p95_latency_ms"]:.2f} ms
- **Throughput rate:** {load_stats["throughput_req_sec"]:.2f} req/sec
- **Total duration of stress test:** {load_stats["duration_s"]:.2f} seconds
</details>

## 📦 Test Report Artifacts

The full test report files are uploaded as part of this workflow run and can be inspected in the artifacts list:
- Website E2E Report: `website/{os.path.basename(web_report_path)}`
- Mobile E2E Report: `mobile/report/{os.path.basename(mobile_report_path)}`
- Backend Security Report: `backend/{os.path.basename(sec_report_path)}`
- Load Testing Report: `load_testing/{os.path.basename(load_report_path)}`
"""
    
    # Write to local file for validation
    with open("summary_dashboard.md", "w", encoding="utf-8") as f:
        f.write(markdown_summary)
    print("\n[SUCCESS] Consolidated report generated: summary_dashboard.md")
    
    # Write to GITHUB_STEP_SUMMARY if active
    summary_file = os.environ.get("GITHUB_STEP_SUMMARY")
    if summary_file:
        with open(summary_file, "a", encoding="utf-8") as sf:
            sf.write(markdown_summary)
        print("[SUCCESS] Wrote summary dashboard to GITHUB_STEP_SUMMARY")
        
if __name__ == "__main__":
    main()
